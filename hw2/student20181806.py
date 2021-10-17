#!/usr/bin/pyton3

from openpyxl import load_workbook

wb = load_workbook( filename = "student.xlsx")
sheet_ranges = wb["Sheet1"]

row_num=2
dic = dict()
while True:
  col_midterm = sheet_ranges.cell( row = row_num, column = 3).value
  col_final = sheet_ranges.cell( row = row_num, column = 4).value
  col_homework = sheet_ranges.cell( row = row_num, column = 5).value
  col_attendance = sheet_ranges.cell(row = row_num, column = 6).value
  if col_midterm == None:
    break
  col_total = col_midterm * 0.3 + col_final * 0.35 + col_homework * 0.34 + col_attendance
  sheet_ranges.cell(row = row_num, column = 7, value = col_total)

  dic[row_num] = col_total

  row_num += 1

num_of_students = row_num - 2
a_plus = (num_of_students * 0.3) * 0.5
a_zero = (num_of_students * 0.3)
b_plus = a_zero + (num_of_students * 0.4 * 0.5)
b_zero = (num_of_students * 0.7)
c_plus = b_zero + ((num_of_students - b_zero) * 0.5)
c_zero = num_of_students

a = a_zero
b = b_zero - a_zero
c = c_zero - b_zero

total = []
for i in dic:
  total.append(dic[i])

rank_result = []
for i in range(0, len(total)):
  r = 0
  for j in range(0, len(total)):
    if total[i] <= total[j]: r += 1
    if total[i] < 40:
      r = -1
  rank_result.append(r)

rank = dict()
i = 0
for s in range(2, row_num):
  rank[s] = rank_result[i]
  i += 1

grades = []
a_check = 0
b_check = 0
c_check = 0

for g in rank:
  if rank[g] == -1:
    grades.append([g, 'F'])
  elif rank[g] < a_plus:
    grades.append([g, 'A+'])
    a_check += 1
  elif rank[g] > a_plus and rank[g] <= a_zero:
    grades.append([g, 'A0'])
  elif rank[g] > a_zero and rank[g] <= b_plus:
    grades.append([g, 'B+'])
    b_check += 1
  elif rank[g] > b_plus and rank[g] <= b_zero:
    grades.append([g, 'B0'])
  elif rank[g] > b_zero and rank[g] <= c_plus:
    grades.append([g, 'C+'])
    c_check += 1
  elif rank[g] > c_plus:
    grades.append([g, 'C0'])

grades = dict(grades)
grades = list(grades.values())

i = 0
if a_check > a / 2 :
  for g in grades:
    if g == 'A+':
      grades[i] = 'A0'
    i += 1

i = 0
if b_check > b / 2:
  for g in grades:
    if g == 'B+':
      grades[i] = 'B0'
    i += 1

i = 0
if c_check > c / 2:
  for g in grades:
    if g == 'C+':
      grades[i] = 'C0'
    i += 1

row_num=2
i = 0
while True:
  col_midterm = sheet_ranges.cell( row = row_num, column = 3).value

  if col_midterm == None:
    break

  col_grade = grades[i]
  sheet_ranges.cell(row = row_num, column = 8, value = col_grade)

  row_num += 1
  i += 1

wb.save( filename = "student.xlsx" )
