from openpyxl import load_workbook

wb = load_workbook( filename = "student.xlsx")
sheet_ranges = wb["Sheet1"]

row_num=2
while True:
	col_midterm = sheet_ranges.cell( row = row_num, column = 3).value
	col_final = sheet_ranges.cell( row = row_num, column = 4).value
	col_homework = sheet_ranges.cell( row = row_num, column = 5).value
	col_attendance = sheet_ranges.cell(row = row_num, column = 6).value
	if col_midterm == None:
		break
	col_total = col_midterm * 0.3 + col_final * 0.35 + col_homework * 0.34 + col_attendance
	sheet_ranges.cell(row = row_num, column = 7, value = col_total)

	#eval_string = "0.3*{}{}+0.35*{}{}+0.34*{}{}+{}{}".format( get_column_letter(3), row_num, 
		#get_column_letter(4), row_num, get_column_letter(5), row_num, get_column_letter(6), row_num )
	#sheet_ranges.cell( row = row_num, column = 7, value = eval_string)
	#print(eval_string)
	row_num += 1

wb.save( filename = "output.xlsx" )
